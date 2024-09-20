import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from data.config import RESULTS_DIR

class ResultProcessor:
    def __init__(self):
        pass

    async def make_table_with_results(self, data):
        result_dict = {}
        for entry in data:
            proxy = entry["proxy"]
            if proxy not in result_dict:
                result_dict[proxy] = {}
            for rpc in entry["rpcs"]:
                result_dict[proxy][rpc["rpc"]] = rpc["status"]

        df = pd.DataFrame(result_dict).T
        df.index.name = "Proxy"

        current_time = datetime.datetime.now()
        timestamp = current_time.strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f"{RESULTS_DIR}rpc_results_{timestamp}.xlsx"
        df.to_excel(file_name)

        wb = load_workbook(file_name)
        ws = wb.active

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                cell.alignment = center_alignment
                if cell.value == True:
                    cell.fill = green_fill
                elif cell.value == "error":
                    cell.fill = red_fill
                elif cell.value == False:
                    cell.fill = red_fill

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_name)